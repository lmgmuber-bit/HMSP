from django.contrib import admin
from django import forms
from .models import Evento, Testimonio, Oracion, Noticia, ConfiguracionSitio, MensajeContacto, Apostolado
from .models import Suscripcion

@admin.register(Suscripcion)
class SuscripcionAdmin(admin.ModelAdmin):
    list_display = ('id', 'nombre', 'telefono', 'email', 'fecha_suscripcion')
    search_fields = ('nombre', 'telefono', 'email')
    ordering = ('-fecha_suscripcion',)

    actions = ['exportar_excel']

    def exportar_excel(self, request, queryset):
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Suscripciones"

        # Encabezados
        headers = ['ID', 'Nombre', 'Teléfono', 'Email', 'Fecha Suscripción']
        ws.append(headers)

        # Datos
        for suscripcion in queryset:
            ws.append([
                suscripcion.id,
                suscripcion.nombre,
                suscripcion.telefono,
                suscripcion.email,
                suscripcion.fecha_suscripcion.strftime('%Y-%m-%d %H:%M')
            ])

        # Ajustar ancho de columnas
        for i, col in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = 20

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=suscripciones.xlsx'
        wb.save(response)
        return response

    exportar_excel.short_description = "Exportar seleccionados a Excel"
from django.core.mail import EmailMultiAlternatives
from django.conf import settings
from .email_templates import render_evento_email, render_noticia_email

@admin.register(Evento)
class EventoAdmin(admin.ModelAdmin):
    list_display = ('titulo', 'fecha', 'orden_carrusel', 'activo', 'fecha_inicio_publicacion', 'fecha_fin_publicacion')
    list_filter = ('activo', 'fecha')
    search_fields = ('titulo', 'descripcion')
    list_editable = ('orden_carrusel', 'activo')
    prepopulated_fields = {'slug': ('titulo',)}
    fieldsets = (
        ('Información Básica', {
            'fields': ('titulo', 'slug', 'descripcion', 'fecha', 'ubicacion', 'imagen'),
            'description': '<strong>📸 ESPECIFICACIONES DE IMAGEN:</strong><br>'
                          '📐 <strong>Dimensiones:</strong> 1200x800 píxeles (carrusel) o 600x400 píxeles (lista)<br>'
                          '📄 <strong>Formato:</strong> JPG (preferido) o PNG<br>'
                          '💾 <strong>Tamaño:</strong> Máximo 2MB<br>'
                          '<em>Tip: Use imágenes horizontales de alta calidad para mejor visualización en el carrusel.</em>'
        }),
        ('Control de Publicación', {
            'fields': ('fecha_inicio_publicacion', 'fecha_fin_publicacion', 'activo', 'orden_carrusel'),
            'description': 'Configure cuándo y cómo se mostrará el evento. El evento se desactivará automáticamente fuera del rango de fechas.'
        }),
        ('Multimedia', {
            'fields': ('tipo_multimedia', 'video_youtube_url', 'video_local', 'preferencia_visualizacion')
        }),
        ('Opciones Adicionales', {
            'fields': ('formulario_url',)
        })
    )
    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        if not change:
            suscriptores = Suscripcion.objects.values_list('email', flat=True)
            if suscriptores:
                    html_content = render_evento_email(obj)
                    subject = f"HMSP - Nuevo evento: {obj.titulo}"
                    from_email = settings.DEFAULT_FROM_EMAIL
                    msg = EmailMultiAlternatives(
                        subject=subject,
                        body=f"Se ha publicado un nuevo evento en HMSP: {obj.titulo}\n\n{obj.descripcion}",
                        from_email=from_email,
                        to=[],
                        bcc=list(suscriptores),
                        headers={"List-Unsubscribe": "<https://hmsp.cl/suscripcion/>"}
                    )
                    msg.attach_alternative(html_content, "text/html")
                    msg.send(fail_silently=True)

@admin.register(Testimonio)
class TestimonioAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'aprobado', 'destacado', 'orden_destacado', 'creado')
    list_filter = ('aprobado', 'destacado', 'creado')
    list_editable = ('destacado', 'orden_destacado')
    search_fields = ('nombre', 'testimonio')
    actions = ['aprobar_testimonios', 'rechazar_testimonios']
    fieldsets = (
        ('Información del Testimonio', {
            'fields': ('nombre', 'testimonio', 'imagen', 'aprobado', 'destacado', 'orden_destacado'),
            'description': '<strong>📸 ESPECIFICACIONES DE IMAGEN:</strong><br>'
                          '📐 <strong>Dimensiones:</strong> 400x400 píxeles (cuadrada 1:1)<br>'
                          '📄 <strong>Formato:</strong> JPG (preferido) o PNG<br>'
                          '💾 <strong>Tamaño:</strong> Máximo 500KB<br>'
                          '<em>Tip: Use fotos de rostro centradas con buena iluminación. Los testimonios deben ser aprobados para aparecer en el sitio.</em><br>'
                          '<strong>⭐ Destacado y Orden:</strong> Marque "Destacado" y asigne orden (1, 2 o 3) para mostrar en la página principal. Use 0 para no destacar.'
        }),
    )

    def aprobar_testimonios(self, request, queryset):
        queryset.update(aprobado=True)
    aprobar_testimonios.short_description = "Aprobar testimonios seleccionados"

    def rechazar_testimonios(self, request, queryset):
        queryset.update(aprobado=False)
    rechazar_testimonios.short_description = "Rechazar testimonios seleccionados"

@admin.register(Oracion)
class OracionAdmin(admin.ModelAdmin):
    list_display = ('titulo', 'categoria', 'activa')
    list_filter = ('categoria', 'activa')
    search_fields = ('titulo', 'contenido')

@admin.register(Noticia)
class NoticiaAdmin(admin.ModelAdmin):
    list_display = ('titulo', 'destacada', 'creado')
    list_filter = ('destacada', 'creado')
    search_fields = ('titulo', 'contenido')
    prepopulated_fields = {'slug': ('titulo',)}
    fieldsets = (
        ('Contenido de la Noticia', {
            'fields': ('titulo', 'slug', 'contenido', 'imagen', 'destacada'),
            'description': '<strong>📸 ESPECIFICACIONES DE IMAGEN:</strong><br>'
                          '📐 <strong>Dimensiones:</strong> 800x600 píxeles (portada) o 400x300 píxeles (miniatura)<br>'
                          '📄 <strong>Formato:</strong> JPG (preferido) o PNG<br>'
                          '💾 <strong>Tamaño:</strong> Máximo 1.5MB<br>'
                          '<em>Tip: Use imágenes bien iluminadas y de alta calidad. Las noticias destacadas aparecen en la página principal.</em>'
        }),
    )
    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        if not change:
            suscriptores = Suscripcion.objects.values_list('email', flat=True)
            if suscriptores:
                    html_content = render_noticia_email(obj)
                    subject = f"HMSP - Nueva noticia: {obj.titulo}"
                    from_email = settings.DEFAULT_FROM_EMAIL
                    msg = EmailMultiAlternatives(
                        subject=subject,
                        body=f"Se ha publicado una nueva noticia en HMSP: {obj.titulo}\n\n{obj.contenido[:200]}...",
                        from_email=from_email,
                        to=[],
                        bcc=list(suscriptores),
                        headers={"List-Unsubscribe": "<https://hmsp.cl/suscripcion/>"}
                    )
                    msg.attach_alternative(html_content, "text/html")
                    msg.send(fail_silently=True)

@admin.register(ConfiguracionSitio)
class ConfiguracionSitioAdmin(admin.ModelAdmin):
    fieldsets = (
        ('Logo del Sitio', {
            'fields': ('logo_principal',),
            'description': '<strong>ESPECIFICACIONES DEL LOGO:</strong><br>'
                          '📐 <strong>Dimensiones:</strong> 200x80 píxeles (recomendado) o máximo 300x120 píxeles<br>'
                          '📄 <strong>Formato:</strong> PNG con fondo transparente (preferido) o JPG<br>'
                          '💾 <strong>Tamaño:</strong> Máximo 500KB<br>'
                          '📱 <strong>Visualización:</strong> Se mostrará con altura de 60px en la navegación<br>'
                          '<em>Tip: Un logo horizontal funciona mejor que uno cuadrado o vertical.</em>'
        }),
        ('Quiénes Somos', {
            'fields': ('quienes_somos_titulo', 'quienes_somos_contenido', 'quienes_somos_imagen'),
            'description': '<strong>Sección Quiénes Somos</strong><br>'
                          '📐 <strong>Imagen:</strong> 1200x600 píxeles | Formato: JPG/PNG | Máximo: 2MB'
        }),
        ('Historia', {
            'fields': ('historia_titulo', 'historia_contenido', 'historia_imagen'),
            'description': '<strong>Sección Historia</strong><br>'
                          '📐 <strong>Imagen:</strong> 800x500 píxeles | Formato: JPG/PNG | Máximo: 1.5MB'
        }),
        ('Vocaciones', {
            'fields': ('vocaciones_titulo', 'vocaciones_contenido', 'vocaciones_imagen'),
            'description': '<strong>Sección Vocaciones</strong><br>'
                          '📐 <strong>Imagen:</strong> 800x500 píxeles | Formato: JPG/PNG | Máximo: 1.5MB'
        }),
        ('Recursos - Información General', {
            'fields': ('recursos_titulo', 'recursos_contenido', 'recursos_imagen'),
            'description': '<strong>Sección Recursos</strong><br>'
                          '📐 <strong>Imagen:</strong> 1200x600 píxeles | Formato: JPG/PNG | Máximo: 2MB<br>'
                          '<em>Contenido introductorio para la página de recursos</em>'
        }),
        ('Recursos - Biblioteca de Oraciones', {
            'fields': ('biblioteca_oraciones_titulo', 'biblioteca_oraciones_contenido', 'biblioteca_oraciones_imagen'),
            'description': '<strong>Subsección: Biblioteca de Oraciones</strong><br>'
                          '📐 <strong>Imagen:</strong> 800x500 píxeles | Formato: JPG/PNG | Máximo: 1.5MB'
        }),
        ('Recursos - Material Espiritual', {
            'fields': ('material_espiritual_titulo', 'material_espiritual_contenido', 'material_espiritual_imagen'),
            'description': '<strong>Subsección: Material Espiritual</strong><br>'
                          '📐 <strong>Imagen:</strong> 800x500 píxeles | Formato: JPG/PNG | Máximo: 1.5MB'
        }),
        ('Recursos - Boletín Mensual', {
            'fields': ('boletin_mensual_titulo', 'boletin_mensual_contenido', 'boletin_mensual_imagen'),
            'description': '<strong>Subsección: Boletín Mensual</strong><br>'
                          '📐 <strong>Imagen:</strong> 800x500 píxeles | Formato: JPG/PNG | Máximo: 1.5MB'
        }),
        ('Recursos - Donaciones', {
            'fields': ('donaciones_titulo', 'donaciones_contenido', 'donaciones_imagen'),
            'description': '<strong>Subsección: Donaciones</strong><br>'
                          '📐 <strong>Imagen:</strong> 800x500 píxeles | Formato: JPG/PNG | Máximo: 1.5MB'
        }),
        ('Información de Contacto', {
            'fields': ('contacto_direccion', 'contacto_telefono', 'contacto_email', 'contacto_horario'),
            'description': 'Datos de contacto que aparecen en el sitio'
        }),
        ('Redes Sociales', {
            'fields': ('facebook_url', 'instagram_url', 'youtube_url', 'twitter_url'),
            'description': 'Enlaces a redes sociales'
        }),
    )
    
    def has_add_permission(self, request):
        # Solo permitir una instancia de configuración
        return not ConfiguracionSitio.objects.exists()
    
    def has_delete_permission(self, request, obj=None):
        # No permitir eliminar la configuración
        return False

@admin.register(MensajeContacto)
class MensajeContactoAdmin(admin.ModelAdmin):
    list_display = ('nombre', 'email', 'asunto', 'leido', 'respondido', 'creado')
    list_filter = ('leido', 'respondido', 'creado')
    list_editable = ('leido', 'respondido')
    search_fields = ('nombre', 'email', 'asunto', 'mensaje')
    readonly_fields = ('nombre', 'email', 'asunto', 'mensaje', 'creado')
    date_hierarchy = 'creado'
    
    fieldsets = (
        ('Información del Contacto', {
            'fields': ('nombre', 'email', 'creado')
        }),
        ('Mensaje', {
            'fields': ('asunto', 'mensaje')
        }),
        ('Estado', {
            'fields': ('leido', 'respondido'),
            'description': 'Marque como leído cuando haya revisado el mensaje. Marque como respondido cuando ya haya contactado a la persona.'
        }),
    )
    
    def has_add_permission(self, request):
        # Los mensajes solo se crean desde el formulario web
        return False


@admin.register(Apostolado)
class ApostoladoAdmin(admin.ModelAdmin):
    list_display = ('titulo', 'orden', 'activo', 'modificado')
    list_filter = ('activo',)
    search_fields = ('titulo', 'descripcion')
    list_editable = ('orden', 'activo')
    prepopulated_fields = {'slug': ('titulo',)}
    fieldsets = (
        ('Información del Apostolado', {
            'fields': ('titulo', 'slug', 'descripcion_corta', 'descripcion', 'imagen', 'orden', 'activo'),
            'description': '<strong>📸 ESPECIFICACIONES DE IMAGEN:</strong><br>'
                          '📐 <strong>Dimensiones:</strong> 800x600px (recomendado)<br>'
                          '📄 <strong>Formato:</strong> JPG, PNG, SVG<br>'
                          '💾 <strong>Tamaño:</strong> Máximo 2MB<br>'
                          '<em>El título aparecerá en el menú. La descripción corta se usa como resumen.</em>'
        }),
    )